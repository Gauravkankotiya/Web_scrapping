import requests
from bs4 import BeautifulSoup
import pandas as pd
import xlsxwriter
import openpyxl
from openpyxl.styles import Font

url = 'https://eprocure.gov.in/cppp/latestactivetendersnew/cpppdata?page=1'


def make_soup(url):
    data = requests.get(url)
    content = data.content
    soup = BeautifulSoup(content, 'html.parser')
    return (soup)


def get_header():
    soup = make_soup(url)
    head = soup.find('th')
    splt = str(f'{head.findNext().findNext().findNext().findNext().text}')
    title, ref, t_id = splt.split('/')
    dcsv= [head.text,head.findNext().text,head.findNext().findNext().text,head.findNext().findNext().findNext().text,title,ref,t_id,head.findNext().findNext().findNext().findNext().findNext().text]
    # print(dcsv)
    #writing in exce; file
    workbook =openpyxl.Workbook()
    worksheet = workbook.active
    for data,i in zip(dcsv,range(1,9)):
        worksheet.cell(1, i).font=Font(size=11,bold=True)
        worksheet.cell(1,i).value=data
    workbook.save('date.xlsx')
    


def get_date(parent):
    head = parent.findChild()
    t_info = head.findNext().findNext().findNext().findNext()
    t_title = str(t_info.find('a').text)  # tender title
    # obtaining ref no. and tender id
    st = str(t_info.text)
    i = st.rfind('/')
    ref_no = st[len(t_title):i]
    t_id = st[i:]

    tcsv = [head.text,head.findNext().text,head.findNext().findNext().text,head.findNext().findNext().findNext().text,t_title,ref_no,t_id,head.findNext().findNext().findNext().findNext().findNext().findNext().text]
    # print(tcsv)
    #writing in excel file
    wb = openpyxl.load_workbook('date.xlsx')
    sheet = wb['Sheet']
    sheet.append(tcsv)
    wb.save('date.xlsx')

    if parent.findNext('tr') != None:
        return get_date(parent.findNext('tr'))

# get_date(soup.find('tr'))


def get_pages():
    
    for i in range(1, 2):
        new_url = url.replace('1', str(i))
        soup = make_soup(new_url)
        get_date(soup.find('tr').findNext('tr'))
        

get_header()
get_pages()
